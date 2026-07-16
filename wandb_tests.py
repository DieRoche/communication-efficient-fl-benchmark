import wandb
import pandas as pd
import json
import re
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── helpers ───────────────────────────────────────────────────────────────────

def fetch_with_retry(fn, *args, max_retries=8, base_wait=30, **kwargs):
    """Retry fn(*args, **kwargs) on HTTP 429 rate-limit with exponential back-off."""
    for attempt in range(max_retries):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            msg = str(e)
            if "429" in msg or "rate limit" in msg.lower():
                wait = base_wait * (2 ** attempt)
                print(f"    Rate limited. Waiting {wait}s before retry {attempt+1}/{max_retries}…")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError(f"Exceeded {max_retries} retries due to rate limiting.")


def safe_filename(name: str, max_len: int = 120) -> str:
    name = name.strip()
    name = re.sub(r"[^\w\-\.]+", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    return name[:max_len] if len(name) > max_len else name


def safe_sheetname(name: str, max_len: int = 31) -> str:
    name = re.sub(r"[\\/*?\[\]:]", "_", name).strip()
    return name[:max_len]


# ── styling ───────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
NORMAL_FONT = Font(name="Arial", size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
THIN        = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data(ws, row, ncols, alternate=False):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        if alternate:
            cell.fill = ALT_FILL
        cell.font = NORMAL_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def autofit(ws, min_w=10, max_w=40):
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, min_w), max_w)


# ── core export ───────────────────────────────────────────────────────────────

def export_by_run_name(entity, project, group, metrics, out_dir):
    """
    For every unique run *name* in the group, create one .xlsx file.
    Each file contains:
      - A 'Summary' sheet: last/mean value per rep + Mean & Std across reps.
      - One sheet per metric: full round-by-round history (one column per rep).
    Histories are fetched once per run to minimise API calls.
    """
    os.makedirs(out_dir, exist_ok=True)

    api  = wandb.Api(timeout=120)          # bumped from default 19 s
    runs = list(fetch_with_retry(api.runs, f"{entity}/{project}", filters={"group": group}))

    # Group by run name
    by_name: dict[str, list] = {}
    for run in runs:
        by_name.setdefault(run.name, []).append(run)

    print(f"Found {len(runs)} runs across {len(by_name)} unique name(s): {list(by_name.keys())}")

    for run_name, run_list in by_name.items():
        print(f"\n── '{run_name}' ({len(run_list)} repetition(s)) ──")
        rep_labels = [f"rep {i+1}" for i in range(len(run_list))]

        # ── Fetch histories ONCE per rep ──────────────────────────────────────
        histories: list = []
        for i, run in enumerate(run_list):
            print(f"  Fetching rep {i+1}/{len(run_list)}  (id: {run.id})…")
            hist = fetch_with_retry(run.history, samples=100_000)
            histories.append(hist if "_step" in hist.columns else pd.DataFrame())
            time.sleep(2)   # polite delay between requests

        # ── Workbook ──────────────────────────────────────────────────────────
        wb = Workbook()
        wb.remove(wb.active)

        # ── Sheet 1: Summary ──────────────────────────────────────────────────
        ws_sum = wb.create_sheet("Summary")
        hdr = ["Metric"] + rep_labels + ["Mean across reps", "Std across reps"]
        for col, val in enumerate(hdr, 1):
            ws_sum.cell(row=1, column=col, value=val)
        style_header(ws_sum, 1, len(hdr))

        n_reps         = len(run_list)
        data_col_start = 2
        data_col_end   = 1 + n_reps

        excel_row = 2
        for stat in ("last", "mean"):
            for m in metrics:
                ws_sum.cell(row=excel_row, column=1, value=f"{m} ({stat})")
                for c_idx, hist in enumerate(histories, 2):
                    val = None
                    if m in hist.columns and not hist.empty:
                        s = pd.to_numeric(hist[m], errors="coerce").dropna()
                        if not s.empty:
                            val = s.iloc[-1] if stat == "last" else s.mean()
                    ws_sum.cell(row=excel_row, column=c_idx, value=val)
                range_str = (
                    f"{get_column_letter(data_col_start)}{excel_row}:"
                    f"{get_column_letter(data_col_end)}{excel_row}"
                )
                ws_sum.cell(row=excel_row, column=data_col_end + 1, value=f'=IFERROR(AVERAGE({range_str}),"")')
                ws_sum.cell(row=excel_row, column=data_col_end + 2, value=f'=IFERROR(STDEV({range_str}),"")')
                style_data(ws_sum, excel_row, len(hdr), alternate=(excel_row % 2 == 0))
                excel_row += 1

        autofit(ws_sum)
        ws_sum.freeze_panes = "B2"

        # ── Sheets 2…N: per-metric history ───────────────────────────────────
        for m in metrics:
            ws = wb.create_sheet(safe_sheetname(m))

            series_list = []
            for i, hist in enumerate(histories):
                if m in hist.columns and not hist.empty:
                    s = (
                        hist[["_step", m]]
                        .copy()
                        .assign(**{m: pd.to_numeric(hist[m], errors="coerce")})
                        .dropna(subset=[m])
                        .drop_duplicates("_step")
                        .set_index("_step")[m]
                    )
                    s.name = rep_labels[i]
                    series_list.append(s)

            if not series_list:
                ws["A1"] = "No data available for this metric."
                continue

            df = pd.concat(series_list, axis=1).sort_index()
            n_rep_cols     = len(df.columns)
            data_col_end_m = 1 + n_rep_cols

            hdr_m = ["round"] + list(df.columns) + ["Mean", "Std"]
            for col, val in enumerate(hdr_m, 1):
                ws.cell(row=1, column=col, value=val)
            style_header(ws, 1, len(hdr_m))

            for r_idx, (step, row_data) in enumerate(df.iterrows()):
                er = r_idx + 2
                ws.cell(row=er, column=1, value=int(step))
                for c_idx, val in enumerate(row_data, 2):
                    ws.cell(row=er, column=c_idx, value=(None if pd.isna(val) else val))
                range_str = (
                    f"{get_column_letter(2)}{er}:"
                    f"{get_column_letter(data_col_end_m)}{er}"
                )
                ws.cell(row=er, column=data_col_end_m + 1, value=f'=IFERROR(AVERAGE({range_str}),"")')
                ws.cell(row=er, column=data_col_end_m + 2, value=f'=IFERROR(STDEV({range_str}),"")')
                style_data(ws, er, len(hdr_m), alternate=(r_idx % 2 == 1))

            autofit(ws)
            ws.freeze_panes = "B2"

        # ── Save ──────────────────────────────────────────────────────────────
        out_path = os.path.join(out_dir, f"{safe_filename(run_name)}.xlsx")
        wb.save(out_path)
        print(f"  Saved → {out_path}")

    print(f"\nDone. All files written to '{out_dir}/'")


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    ENTITY = os.environ.get("WANDB_ENTITY")
    if not ENTITY:
        raise SystemExit("Set WANDB_ENTITY explicitly before exporting W&B runs.")
    PROJECT = os.environ.get("WANDB_PROJECT", "communication-efficient-fl-benchmark")
    GROUP = os.environ.get("WANDB_GROUP")
    if not GROUP:
        raise SystemExit("Set WANDB_GROUP explicitly before exporting W&B runs.")

    out_dir = f"{PROJECT}_wandb"

    METRICS = [
        "acc_servers_highest",
               

        "round_flops",
        "serialization_flops",
        "decompression_flops_server",
        "decompression_flops_clients",
        "compression_flops_server",
        "compression_flops_clients",


        "overall_traffic",
        "download_traffic",
        "upload_traffic",
        "upload_traffic_per_client",

        "training_flops",
        "evaluation_flops",
        "aggregation_flops",
    ]

    export_by_run_name(ENTITY, PROJECT, GROUP, METRICS, out_dir)
